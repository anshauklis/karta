import io
import api.json_util as json
import secrets
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import text

from api.database import engine
from api.auth.dependencies import get_current_user, check_ownership, require_role
from api.models import SharedLinkCreate, SharedLinkResponse, ImportConfirmRequest, ExcelExportRequest

router = APIRouter(tags=["export"])


def _get_excel_format(fmt: dict) -> str | None:
    """Convert column format config to Excel number format string."""
    fmt_type = fmt.get("type")
    if not fmt_type:
        return None
    decimals = fmt.get("decimals", 0) or 0
    dec_part = "." + "0" * decimals if decimals > 0 else ""

    if fmt_type == "number":
        return f"#,##0{dec_part}" if fmt.get("thousands", True) else f"0{dec_part}"
    if fmt_type == "currency":
        sym = fmt.get("prefix") or "$"
        return f"{sym}#,##0{dec_part}" if fmt.get("thousands", True) else f"{sym}0{dec_part}"
    if fmt_type == "percent":
        return f"0{dec_part}%"
    if fmt_type == "date":
        pattern = fmt.get("date_pattern", "YYYY-MM-DD")
        date_map = {
            "DD.MM.YYYY": "DD.MM.YYYY",
            "MM/DD/YYYY": "MM/DD/YYYY",
            "DD Mon YYYY": "DD MMM YYYY",
        }
        return date_map.get(pattern, "YYYY-MM-DD")
    return None


@router.post("/api/export/xlsx", summary="Generate Excel file from data")
def export_xlsx(body: ExcelExportRequest, current_user: dict = Depends(get_current_user)):
    """Accept columns + rows + formatting config, return .xlsx file download."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    columns = body.columns
    rows = body.rows

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for ri, row in enumerate(rows, 2):
        for ci, value in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=value)

    # Column number formats
    if body.column_formats:
        for ci, col_name in enumerate(columns):
            fmt = body.column_formats.get(col_name)
            if not fmt or not isinstance(fmt, dict):
                continue
            excel_fmt = _get_excel_format(fmt)
            if excel_fmt:
                for ri in range(2, len(rows) + 2):
                    ws.cell(row=ri, column=ci + 1).number_format = excel_fmt

    # Conditional formatting (threshold rules → cell fill/font)
    if body.formatting:
        # Index rules by column
        rules_by_col: dict[str, list[dict]] = {}
        for rule in body.formatting:
            cols = rule.get("columns") or ([rule["column"]] if rule.get("column") else [])
            for col in cols:
                rules_by_col.setdefault(col, []).append(rule)

        for ci, col_name in enumerate(columns):
            col_rules = rules_by_col.get(col_name)
            if not col_rules:
                continue
            for ri, row in enumerate(rows):
                value = row[ci] if ci < len(row) else None
                try:
                    num_val = float(value) if value is not None else None
                except (ValueError, TypeError):
                    continue
                if num_val is None:
                    continue
                for rule in col_rules:
                    if rule.get("type") != "threshold" or not rule.get("rules"):
                        continue
                    for r in rule["rules"]:
                        op = r.get("op", "")
                        threshold = r.get("value", 0)
                        match = False
                        if op == ">" and num_val > threshold:
                            match = True
                        elif op == "<" and num_val < threshold:
                            match = True
                        elif op == ">=" and num_val >= threshold:
                            match = True
                        elif op == "<=" and num_val <= threshold:
                            match = True
                        elif op == "=" and num_val == threshold:
                            match = True
                        elif op == "!=" and num_val != threshold:
                            match = True
                        if match:
                            cell = ws.cell(row=ri + 2, column=ci + 1)
                            color = r.get("color", "").lstrip("#")
                            if color:
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            text_color = r.get("text_color", "").lstrip("#")
                            if text_color:
                                cell.font = Font(color=text_color)
                            break

    # Auto-width columns
    for ci, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for ri in range(min(len(rows), 100)):
            val = rows[ri][ci - 1] if (ci - 1) < len(rows[ri]) else None
            val_len = len(str(val)) if val is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 3, 40)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = body.filename
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/dashboards/{dashboard_id}/export", summary="Export dashboard as JSON")
def export_dashboard(dashboard_id: int, current_user: dict = Depends(get_current_user)):
    """Export a dashboard with all tabs, connections, datasets, charts, and filters as a downloadable JSON file."""
    with engine.connect() as conn:
        # 1. Dashboard
        dash = conn.execute(text("""
            SELECT id, title, description, icon, url_slug, color_scheme, filter_layout
            FROM dashboards WHERE id = :id
        """), {"id": dashboard_id}).mappings().first()
        if not dash:
            raise HTTPException(404, "Dashboard not found")
        dash = dict(dash)

        # Roles
        roles_rows = conn.execute(text(
            "SELECT group_name FROM dashboard_roles WHERE dashboard_id = :did"
        ), {"did": dashboard_id}).all()
        roles = [r[0] for r in roles_rows]

        dashboard_export = {
            "title": dash["title"],
            "description": dash["description"],
            "icon": dash["icon"],
            "url_slug": dash["url_slug"],
            "color_scheme": dash["color_scheme"],
            "filter_layout": dash["filter_layout"],
            "roles": roles,
        }

        # 2. Tabs (ordered by position_order, assign _ref)
        tabs_rows = conn.execute(text(
            "SELECT id, title, position_order FROM dashboard_tabs WHERE dashboard_id = :did ORDER BY position_order"
        ), {"did": dashboard_id}).mappings().all()

        tab_id_to_ref = {}
        tabs_export = []
        for i, tab in enumerate(tabs_rows):
            ref = f"tab_{i}"
            tab_id_to_ref[tab["id"]] = ref
            tabs_export.append({
                "_ref": ref,
                "title": tab["title"],
                "position_order": tab["position_order"],
            })

        # 3. Charts (need IDs for filter scope remapping)
        charts_rows = conn.execute(text("""
            SELECT id, connection_id, dataset_id, title, description, mode,
                   chart_type, chart_config, chart_code, sql_query, position_order,
                   grid_x, grid_y, grid_w, grid_h, tab_id
            FROM charts WHERE dashboard_id = :did ORDER BY position_order
        """), {"did": dashboard_id}).mappings().all()

        # Collect all unique connection_ids and dataset_ids used by charts
        connection_ids = set()
        dataset_ids = set()
        for chart in charts_rows:
            if chart["connection_id"]:
                connection_ids.add(chart["connection_id"])
            if chart["dataset_id"]:
                dataset_ids.add(chart["dataset_id"])

        # Also collect dataset_ids referenced by filters
        filters_rows = conn.execute(text("""
            SELECT id, label, filter_type, target_column, default_value, sort_order, config, group_name
            FROM dashboard_filters WHERE dashboard_id = :did ORDER BY sort_order
        """), {"did": dashboard_id}).mappings().all()

        for f in filters_rows:
            cfg = f["config"] or {}
            if cfg.get("dataset_id"):
                dataset_ids.add(cfg["dataset_id"])

        # Datasets may reference additional connections
        if dataset_ids:
            ds_conn_rows = conn.execute(text(
                "SELECT DISTINCT connection_id FROM datasets WHERE id = ANY(:ids) AND connection_id IS NOT NULL"
            ), {"ids": list(dataset_ids)}).all()
            for row in ds_conn_rows:
                connection_ids.add(row[0])

        # 4. Connections (NEVER export passwords)
        conn_id_to_ref = {}
        connections_export = []
        if connection_ids:
            conn_rows = conn.execute(text("""
                SELECT id, name, db_type, host, port, database_name, username, ssl_enabled
                FROM connections WHERE id = ANY(:ids) ORDER BY id
            """), {"ids": list(connection_ids)}).mappings().all()
            for i, c in enumerate(conn_rows):
                ref = f"conn_{i}"
                conn_id_to_ref[c["id"]] = ref
                connections_export.append({
                    "_ref": ref,
                    "name": c["name"],
                    "db_type": c["db_type"],
                    "host": c["host"],
                    "port": c["port"],
                    "database_name": c["database_name"],
                    "username": c["username"],
                    "ssl_enabled": c["ssl_enabled"],
                })

        # 5. Datasets
        ds_id_to_ref = {}
        datasets_export = []
        if dataset_ids:
            ds_rows = conn.execute(text("""
                SELECT id, connection_id, name, description, sql_query, cache_ttl,
                       dataset_type, table_name, schema_name
                FROM datasets WHERE id = ANY(:ids) ORDER BY id
            """), {"ids": list(dataset_ids)}).mappings().all()
            for i, ds in enumerate(ds_rows):
                ref = f"ds_{i}"
                ds_id_to_ref[ds["id"]] = ref
                datasets_export.append({
                    "_ref": ref,
                    "_connection_ref": conn_id_to_ref.get(ds["connection_id"]),
                    "name": ds["name"],
                    "description": ds["description"],
                    "sql_query": ds["sql_query"],
                    "cache_ttl": ds["cache_ttl"],
                    "dataset_type": ds["dataset_type"],
                    "table_name": ds["table_name"],
                    "schema_name": ds["schema_name"],
                })

        # 6. Charts export (with ref mappings)
        chart_id_to_index = {}
        charts_export = []
        for i, chart in enumerate(charts_rows):
            chart_id_to_index[chart["id"]] = i
            charts_export.append({
                "_tab_ref": tab_id_to_ref.get(chart["tab_id"]),
                "_connection_ref": conn_id_to_ref.get(chart["connection_id"]),
                "_dataset_ref": ds_id_to_ref.get(chart["dataset_id"]),
                "title": chart["title"],
                "description": chart["description"],
                "mode": chart["mode"],
                "chart_type": chart["chart_type"],
                "chart_config": chart["chart_config"],
                "chart_code": chart["chart_code"],
                "sql_query": chart["sql_query"],
                "grid_x": chart["grid_x"],
                "grid_y": chart["grid_y"],
                "grid_w": chart["grid_w"],
                "grid_h": chart["grid_h"],
                "position_order": chart["position_order"],
            })

        # 7. Filters export (remap config references)
        filters_export = []
        for f in filters_rows:
            cfg = dict(f["config"]) if f["config"] else {}

            # Remap dataset_id -> _dataset_ref
            original_ds_id = cfg.pop("dataset_id", None)
            if original_ds_id:
                cfg["_dataset_ref"] = ds_id_to_ref.get(original_ds_id)

            # Remap scope: {chart_id: column_name} -> {_chart_index_N: column_name}
            original_scope = cfg.pop("scope", None)
            if original_scope and isinstance(original_scope, dict):
                remapped_scope = {}
                for chart_id_str, col_name in original_scope.items():
                    try:
                        cid = int(chart_id_str)
                    except (ValueError, TypeError):
                        continue
                    idx = chart_id_to_index.get(cid)
                    if idx is not None:
                        remapped_scope[f"_chart_index_{idx}"] = col_name
                cfg["scope"] = remapped_scope

            # Remove depends_on_filter_id (too complex to remap)
            cfg.pop("depends_on_filter_id", None)

            filters_export.append({
                "label": f["label"],
                "filter_type": f["filter_type"],
                "target_column": f["target_column"],
                "default_value": f["default_value"],
                "sort_order": f["sort_order"],
                "config": cfg,
                "group_name": f["group_name"],
            })

        # Build final export payload
        payload = {
            "version": 1,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "source": "Karta",
            "dashboard": dashboard_export,
            "tabs": tabs_export,
            "connections": connections_export,
            "datasets": datasets_export,
            "charts": charts_export,
            "filters": filters_export,
        }

        filename = f"{dash['url_slug']}.json"
        return Response(
            content=json.dumps(payload, ensure_ascii=False, indent=2, default=str),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@router.post("/api/dashboards/import", summary="Preview dashboard import")
def import_preview(body: dict, current_user: dict = require_role("editor", "admin")):
    """Accept an export JSON, validate it, auto-match connections, and return a preview for confirmation."""
    # 1. Validate structure
    if body.get("version") != 1:
        raise HTTPException(400, "Unsupported export version (expected 1)")
    if "dashboard" not in body:
        raise HTTPException(400, "Missing 'dashboard' key in export data")
    if "charts" not in body:
        raise HTTPException(400, "Missing 'charts' key in export data")

    exported_dashboard = body["dashboard"]
    exported_connections = body.get("connections", [])
    exported_charts = body.get("charts", [])
    exported_tabs = body.get("tabs", [])
    exported_filters = body.get("filters", [])

    # 2. Get all connections on this server
    with engine.connect() as conn:
        server_conns = conn.execute(text(
            "SELECT id, name, db_type, host, database_name FROM connections ORDER BY name"
        )).mappings().all()

    # Build lookup structures for matching
    by_name: dict[str, dict] = {}
    by_fingerprint: dict[str, dict] = {}
    available_connections = []
    for sc in server_conns:
        sc_dict = dict(sc)
        available_connections.append({"id": sc_dict["id"], "name": sc_dict["name"], "db_type": sc_dict["db_type"]})
        by_name[sc_dict["name"]] = sc_dict
        fp = f"{sc_dict['db_type']}:{sc_dict['host']}:{sc_dict['database_name']}"
        by_fingerprint[fp] = sc_dict

    # 3. Match each exported connection
    connections_preview = []
    for ec in exported_connections:
        ref = ec.get("_ref", "")
        exported_info = {
            "name": ec.get("name", ""),
            "db_type": ec.get("db_type", ""),
            "host": ec.get("host", ""),
            "database_name": ec.get("database_name", ""),
        }

        matched = None
        # Try exact name match first
        if ec.get("name") and ec["name"] in by_name:
            matched = by_name[ec["name"]]
        # Fallback: fingerprint match (db_type + host + database_name)
        if not matched:
            fp = f"{ec.get('db_type', '')}:{ec.get('host', '')}:{ec.get('database_name', '')}"
            if fp in by_fingerprint:
                matched = by_fingerprint[fp]

        connections_preview.append({
            "_ref": ref,
            "exported": exported_info,
            "status": "matched" if matched else "unmatched",
            "matched_connection_id": matched["id"] if matched else None,
            "matched_connection_name": matched["name"] if matched else None,
        })

    # 4. Build dashboard summary
    dashboard_preview = {
        "title": exported_dashboard.get("title", ""),
        "description": exported_dashboard.get("description", ""),
        "icon": exported_dashboard.get("icon", "📊"),
        "chart_count": len(exported_charts),
        "tab_count": len(exported_tabs),
        "filter_count": len(exported_filters),
    }

    return {
        "dashboard": dashboard_preview,
        "connections": connections_preview,
        "available_connections": available_connections,
    }


@router.post("/api/dashboards/import/confirm", summary="Confirm dashboard import")
def import_confirm(body: ImportConfirmRequest, current_user: dict = require_role("editor", "admin")):
    """Accept confirmed import with connection mappings. Creates dashboard with all entities."""
    user_id = int(current_user["sub"])
    data = body.data
    connection_mapping = body.connection_mapping  # e.g. {"conn_0": 5, "conn_1": 12}

    # Validate structure
    if data.get("version") != 1:
        raise HTTPException(400, "Unsupported export version (expected 1)")
    if "dashboard" not in data:
        raise HTTPException(400, "Missing 'dashboard' key in export data")

    exported_dashboard = data["dashboard"]
    exported_tabs = data.get("tabs", [])
    exported_datasets = data.get("datasets", [])
    exported_charts = data.get("charts", [])
    exported_filters = data.get("filters", [])

    with engine.connect() as conn:
        # 1. Resolve unique slug
        base_slug = exported_dashboard.get("url_slug", "imported-dashboard") or "imported-dashboard"
        slug = base_slug
        suffix = 2
        while True:
            exists = conn.execute(
                text("SELECT COUNT(*) FROM dashboards WHERE url_slug = :slug"),
                {"slug": slug}
            ).scalar()
            if exists == 0:
                break
            slug = f"{base_slug}-{suffix}"
            suffix += 1

        # 2. Create dashboard
        max_order = conn.execute(text("SELECT COALESCE(MAX(sort_order), -1) FROM dashboards")).scalar()
        filter_layout = exported_dashboard.get("filter_layout")
        filter_layout_json = json.dumps(filter_layout) if filter_layout else "{}"

        new_dash = conn.execute(text("""
            INSERT INTO dashboards (title, description, icon, url_slug, sort_order, created_by, filter_layout, color_scheme)
            VALUES (:title, :desc, :icon, :slug, :sort_order, :uid, CAST(:filter_layout AS jsonb), :color_scheme)
            RETURNING id
        """), {
            "title": exported_dashboard.get("title", "Imported Dashboard"),
            "desc": exported_dashboard.get("description", ""),
            "icon": exported_dashboard.get("icon", "\U0001f4ca"),
            "slug": slug,
            "sort_order": max_order + 1,
            "uid": user_id,
            "filter_layout": filter_layout_json,
            "color_scheme": exported_dashboard.get("color_scheme"),
        }).mappings().fetchone()
        new_dash_id = new_dash["id"]

        # 3. Add creator as owner
        conn.execute(text(
            "INSERT INTO dashboard_owners (dashboard_id, user_id) VALUES (:did, :uid)"
        ), {"did": new_dash_id, "uid": user_id})

        # 4. Add roles
        for role in exported_dashboard.get("roles", []):
            if role:
                conn.execute(text(
                    "INSERT INTO dashboard_roles (dashboard_id, group_name) VALUES (:did, :group)"
                ), {"did": new_dash_id, "group": role})

        # 5. Create tabs -> build tab_ref_to_id map
        tab_ref_to_id = {}
        if exported_tabs:
            for tab in exported_tabs:
                new_tab = conn.execute(text("""
                    INSERT INTO dashboard_tabs (dashboard_id, title, position_order)
                    VALUES (:did, :title, :pos) RETURNING id
                """), {
                    "did": new_dash_id,
                    "title": tab.get("title", "Tab"),
                    "pos": tab.get("position_order", 0),
                }).mappings().fetchone()
                tab_ref_to_id[tab.get("_ref", "")] = new_tab["id"]
        else:
            # No tabs in export — create a default "Main" tab
            default_tab = conn.execute(text("""
                INSERT INTO dashboard_tabs (dashboard_id, title, position_order)
                VALUES (:did, 'Main', 0) RETURNING id
            """), {"did": new_dash_id}).mappings().fetchone()
            tab_ref_to_id["_default"] = default_tab["id"]

        # 6. Resolve datasets -> build dataset_ref_to_id map
        dataset_ref_to_id = {}
        for ds in exported_datasets:
            ds_ref = ds.get("_ref", "")
            conn_ref = ds.get("_connection_ref")
            ds_connection_id = connection_mapping.get(conn_ref) if conn_ref else None

            ds_name = ds.get("name", "")

            # Check if dataset with same name AND same connection_id already exists
            existing_ds = None
            if ds_name and ds_connection_id is not None:
                existing_ds = conn.execute(text(
                    "SELECT id FROM datasets WHERE name = :name AND connection_id = :cid"
                ), {"name": ds_name, "cid": ds_connection_id}).mappings().first()

            if existing_ds:
                dataset_ref_to_id[ds_ref] = existing_ds["id"]
            else:
                # Create new dataset
                new_ds = conn.execute(text("""
                    INSERT INTO datasets (connection_id, name, description, sql_query, cache_ttl,
                                          dataset_type, table_name, schema_name, created_by)
                    VALUES (:cid, :name, :desc, :sql, :ttl, :dtype, :tname, :sname, :uid)
                    RETURNING id
                """), {
                    "cid": ds_connection_id,
                    "name": ds_name,
                    "desc": ds.get("description", ""),
                    "sql": ds.get("sql_query", ""),
                    "ttl": ds.get("cache_ttl", 600),
                    "dtype": ds.get("dataset_type", "virtual"),
                    "tname": ds.get("table_name"),
                    "sname": ds.get("schema_name"),
                    "uid": user_id,
                }).mappings().fetchone()
                dataset_ref_to_id[ds_ref] = new_ds["id"]

        # 7. Create charts -> build chart_index_to_id map
        chart_index_to_id = {}
        for i, chart in enumerate(exported_charts):
            # Resolve refs
            tab_ref = chart.get("_tab_ref")
            tab_id = tab_ref_to_id.get(tab_ref) if tab_ref else None
            # If no tab_id resolved but we have a default tab, use it
            if tab_id is None and "_default" in tab_ref_to_id:
                tab_id = tab_ref_to_id["_default"]

            conn_ref = chart.get("_connection_ref")
            chart_connection_id = connection_mapping.get(conn_ref) if conn_ref else None

            ds_ref = chart.get("_dataset_ref")
            chart_dataset_id = dataset_ref_to_id.get(ds_ref) if ds_ref else None

            config_json = json.dumps(chart.get("chart_config")) if chart.get("chart_config") else "{}"

            new_chart = conn.execute(text("""
                INSERT INTO charts (dashboard_id, connection_id, dataset_id, title, description, mode,
                    chart_type, chart_config, chart_code, sql_query, position_order,
                    grid_x, grid_y, grid_w, grid_h, tab_id, created_by)
                VALUES (:did, :cid, :dsid, :title, :desc, :mode,
                    :ctype, CAST(:config AS jsonb), :code, :sql, :pos,
                    :gx, :gy, :gw, :gh, :tid, :uid)
                RETURNING id
            """), {
                "did": new_dash_id,
                "cid": chart_connection_id,
                "dsid": chart_dataset_id,
                "title": chart.get("title", ""),
                "desc": chart.get("description", ""),
                "mode": chart.get("mode", "visual"),
                "ctype": chart.get("chart_type", "bar"),
                "config": config_json,
                "code": chart.get("chart_code"),
                "sql": chart.get("sql_query"),
                "pos": chart.get("position_order", i),
                "gx": chart.get("grid_x", 0),
                "gy": chart.get("grid_y", 0),
                "gw": chart.get("grid_w", 6),
                "gh": chart.get("grid_h", 224),
                "tid": tab_id,
                "uid": user_id,
            }).mappings().fetchone()
            chart_index_to_id[i] = new_chart["id"]

        # 8. Create filters (remap dataset_ref and chart scope)
        for f in exported_filters:
            cfg = dict(f.get("config", {})) if f.get("config") else {}

            # Remap _dataset_ref -> dataset_id
            ds_ref = cfg.pop("_dataset_ref", None)
            if ds_ref:
                cfg["dataset_id"] = dataset_ref_to_id.get(ds_ref)

            # Remap scope: {"_chart_index_0": "col"} -> {new_chart_id: "col"}
            portable_scope = cfg.pop("scope", None)
            if portable_scope and isinstance(portable_scope, dict):
                real_scope = {}
                for key, col_name in portable_scope.items():
                    # key is like "_chart_index_0"
                    if key.startswith("_chart_index_"):
                        try:
                            idx = int(key.replace("_chart_index_", ""))
                        except (ValueError, TypeError):
                            continue
                        new_chart_id = chart_index_to_id.get(idx)
                        if new_chart_id is not None:
                            real_scope[str(new_chart_id)] = col_name
                cfg["scope"] = real_scope

            filter_config_json = json.dumps(cfg)
            conn.execute(text("""
                INSERT INTO dashboard_filters (dashboard_id, label, filter_type, target_column,
                    default_value, sort_order, config, group_name)
                VALUES (:did, :label, :ftype, :col, :default, :order, CAST(:config AS jsonb), :group)
            """), {
                "did": new_dash_id,
                "label": f.get("label", ""),
                "ftype": f.get("filter_type", "select"),
                "col": f.get("target_column", ""),
                "default": f.get("default_value"),
                "order": f.get("sort_order", 0),
                "config": filter_config_json,
                "group": f.get("group_name"),
            })

        # 9. Commit transaction
        conn.commit()

    # 10. Return new dashboard info
    return {"id": new_dash_id, "url_slug": slug}


@router.post("/api/dashboards/{dashboard_id}/share", response_model=SharedLinkResponse, summary="Create share link")
def create_share_link(dashboard_id: int, body: SharedLinkCreate, current_user: dict = require_role("editor", "admin")):
    """Generate a unique share token for a dashboard with optional expiration."""
    uid = int(current_user["sub"])
    token = secrets.token_urlsafe(32)
    expires_at = None
    if body.expires_in_hours:
        expires_at = datetime.now(timezone.utc) + timedelta(hours=body.expires_in_hours)

    with engine.connect() as conn:
        # Verify dashboard exists
        dash = conn.execute(text("SELECT id FROM dashboards WHERE id = :id"), {"id": dashboard_id}).first()
        if not dash:
            raise HTTPException(404, "Dashboard not found")

        row = conn.execute(text("""
            INSERT INTO shared_links (dashboard_id, token, created_by, expires_at)
            VALUES (:did, :token, :uid, :expires)
            RETURNING *
        """), {"did": dashboard_id, "token": token, "uid": uid, "expires": expires_at})
        conn.commit()
        return dict(row.mappings().first())


@router.get("/api/dashboards/{dashboard_id}/shares", response_model=list[SharedLinkResponse], summary="List share links")
def list_share_links(dashboard_id: int, current_user: dict = Depends(get_current_user)):
    """Return all share links for a dashboard, most recent first."""
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT * FROM shared_links
            WHERE dashboard_id = :did
            ORDER BY created_at DESC
        """), {"did": dashboard_id})
        return [dict(r) for r in rows.mappings().all()]


@router.delete("/api/shares/{link_id}", summary="Delete share link")
def revoke_share_link(link_id: int, current_user: dict = require_role("editor", "admin")):
    """Revoke and delete a share link so it can no longer be used."""
    with engine.connect() as conn:
        check_ownership(conn, "shared_links", link_id, current_user)
        conn.execute(text("DELETE FROM shared_links WHERE id = :id"), {"id": link_id})
        conn.commit()
    return {"ok": True}


@router.post("/api/charts/{chart_id}/share", response_model=SharedLinkResponse, summary="Create chart share link")
def create_chart_share_link(chart_id: int, body: SharedLinkCreate, current_user: dict = require_role("editor", "admin")):
    """Generate a unique share token for a single chart."""
    uid = int(current_user["sub"])
    token = secrets.token_urlsafe(32)
    expires_at = None
    if body.expires_in_hours:
        expires_at = datetime.now(timezone.utc) + timedelta(hours=body.expires_in_hours)

    with engine.connect() as conn:
        chart = conn.execute(text("SELECT id, dashboard_id FROM charts WHERE id = :id"), {"id": chart_id}).mappings().first()
        if not chart:
            raise HTTPException(404, "Chart not found")

        row = conn.execute(text("""
            INSERT INTO shared_links (dashboard_id, chart_id, token, created_by, expires_at)
            VALUES (:did, :cid, :token, :uid, :expires)
            RETURNING *
        """), {"did": chart["dashboard_id"], "cid": chart_id, "token": token, "uid": uid, "expires": expires_at})
        conn.commit()
        return dict(row.mappings().first())


@router.get("/api/shared/chart/{token}", summary="Get shared chart")
def get_shared_chart(token: str, filters: str | None = Query(None)):
    """Public endpoint (no auth). Resolve a chart share token and return the chart with executed data."""
    from api.charts.router import _execute_chart_full
    from api.executor import build_visual_chart, build_pivot_table, execute_chart_code

    with engine.connect() as conn:
        link = conn.execute(text("""
            SELECT * FROM shared_links WHERE token = :token AND chart_id IS NOT NULL
        """), {"token": token}).mappings().first()

        if not link:
            raise HTTPException(404, "Share link not found")

        if link["expires_at"] and link["expires_at"] < datetime.now(timezone.utc):
            raise HTTPException(410, "Share link has expired")

        chart = conn.execute(text("""
            SELECT id, title, description, chart_type, chart_config, sql_query,
                   connection_id, mode, chart_code, created_by, created_at, updated_at
            FROM charts WHERE id = :id
        """), {"id": link["chart_id"]}).mappings().first()

        if not chart:
            raise HTTPException(404, "Chart not found")

        # Parse embed filters
        parsed_filters = None
        if filters:
            try:
                parsed_filters = json.loads(filters)
                if not isinstance(parsed_filters, dict):
                    parsed_filters = None
            except (ValueError, TypeError):
                parsed_filters = None

        share_creator_id = link["created_by"]
        chart_dict = dict(chart)
        result = {"figure": None, "columns": [], "rows": [], "row_count": 0, "error": None, "formatting": []}

        try:
            if chart["sql_query"] and chart["connection_id"]:
                chart_config = chart["chart_config"] or {}
                columns, rows, df, pq_path = _execute_chart_full(
                    chart["connection_id"], chart["sql_query"], chart_config,
                    filters=parsed_filters,
                    user_id=share_creator_id)
                figure = None
                if chart["mode"] == "visual" and chart["chart_type"] == "pivot":
                    pivot_result = build_pivot_table(chart_config, df)
                    result = {"figure": None, "columns": pivot_result["columns"], "rows": pivot_result["rows"][:500], "row_count": pivot_result["row_count"], "error": None, "formatting": pivot_result.get("formatting", [])}
                else:
                    if chart["mode"] == "visual":
                        figure = build_visual_chart(chart["chart_type"], chart_config, df)
                    elif chart["mode"] == "code":
                        figure = execute_chart_code(chart["chart_code"], df, parquet_path=pq_path)
                    formatting = chart_config.get("conditional_formatting", []) if chart_config else []
                    result = {"figure": figure, "columns": columns, "rows": [list(r) for r in rows[:200]], "row_count": len(rows), "error": None, "formatting": formatting}
        except Exception as e:
            result["error"] = str(e)

        chart_dict["result"] = result
        return {"chart": chart_dict}


@router.get("/api/charts/{chart_id}/shares", response_model=list[SharedLinkResponse], summary="List chart share links")
def list_chart_share_links(chart_id: int, current_user: dict = Depends(get_current_user)):
    """Return all share links for a chart."""
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT * FROM shared_links
            WHERE chart_id = :cid
            ORDER BY created_at DESC
        """), {"cid": chart_id})
        return [dict(r) for r in rows.mappings().all()]


@router.get("/api/shared/{token}", summary="Get shared dashboard")
def get_shared_dashboard(token: str, filters: str | None = Query(None)):
    """Public endpoint (no auth). Resolve a share token and return the dashboard with executed charts."""
    from api.charts.router import _execute_chart_full
    from api.executor import build_visual_chart, build_pivot_table, execute_chart_code

    with engine.connect() as conn:
        link = conn.execute(text("""
            SELECT * FROM shared_links WHERE token = :token
        """), {"token": token}).mappings().first()

        if not link:
            raise HTTPException(404, "Share link not found")

        if link["expires_at"] and link["expires_at"] < datetime.now(timezone.utc):
            raise HTTPException(410, "Share link has expired")

        dashboard = conn.execute(text("""
            SELECT * FROM dashboards WHERE id = :id
        """), {"id": link["dashboard_id"]}).mappings().first()

        if not dashboard:
            raise HTTPException(404, "Dashboard not found")

        charts = conn.execute(text("""
            SELECT id, title, description, chart_type, chart_config, sql_query,
                   connection_id, mode, chart_code,
                   COALESCE(grid_x, 0) as grid_x, COALESCE(grid_y, 0) as grid_y,
                   COALESCE(grid_w, 6) as grid_w, COALESCE(grid_h, 224) as grid_h,
                   position_order, created_by, created_at, updated_at
            FROM charts WHERE dashboard_id = :did ORDER BY position_order
        """), {"did": dashboard["id"]}).mappings().all()

        # Apply RLS using the share link creator's identity
        share_creator_id = link["created_by"]

        # Parse embed filters
        parsed_filters = None
        if filters:
            try:
                parsed_filters = json.loads(filters)
                if not isinstance(parsed_filters, dict):
                    parsed_filters = None
            except (ValueError, TypeError):
                parsed_filters = None

        chart_results = []
        for chart in charts:
            chart_dict = dict(chart)
            result = {"figure": None, "columns": [], "rows": [], "row_count": 0, "error": None, "formatting": []}
            try:
                if chart["sql_query"] and chart["connection_id"]:
                    chart_config = chart["chart_config"] or {}
                    columns, rows, df, pq_path = _execute_chart_full(
                        chart["connection_id"], chart["sql_query"], chart_config,
                        filters=parsed_filters,
                        user_id=share_creator_id)
                    figure = None
                    if chart["mode"] == "visual" and chart["chart_type"] == "pivot":
                        pivot_result = build_pivot_table(chart_config, df)
                        result = {"figure": None, "columns": pivot_result["columns"], "rows": pivot_result["rows"][:500], "row_count": pivot_result["row_count"], "error": None, "formatting": pivot_result.get("formatting", [])}
                    else:
                        if chart["mode"] == "visual":
                            figure = build_visual_chart(chart["chart_type"], chart_config, df)
                        elif chart["mode"] == "code":
                            figure = execute_chart_code(chart["chart_code"], df, parquet_path=pq_path)
                        formatting = chart_config.get("conditional_formatting", []) if chart_config else []
                        result = {"figure": figure, "columns": columns, "rows": [list(r) for r in rows[:200]], "row_count": len(rows), "error": None, "formatting": formatting}
            except Exception as e:
                result["error"] = str(e)
            chart_dict["result"] = result
            chart_results.append(chart_dict)

        return {
            "dashboard": dict(dashboard),
            "charts": chart_results,
        }
